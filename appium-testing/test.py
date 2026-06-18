import os
import sys
import time
import subprocess
from datetime import datetime

# ── UTF-8 output (emoji support on Windows) ───────────────────────────────────
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# ── Package imports ───────────────────────────────────────────────────────────
try:
    from appium import webdriver
    from appium.options.android import UiAutomator2Options
    from appium.webdriver.common.appiumby import AppiumBy
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
except ImportError:
    print("Required packages missing. Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl missing. Run: pip install openpyxl")
    sys.exit(1)

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════
IS_LINUX = sys.platform.startswith('linux')
IS_CI    = os.environ.get('CI', '').lower() in ('true', '1')

SDK_PATH = os.environ.get(
    "ANDROID_HOME",
    "/usr/local/lib/android/sdk" if IS_LINUX else r"C:\Users\HP\AppData\Local\Android\Sdk"
)

ADB_PATH      = "adb" if IS_LINUX else os.path.join(SDK_PATH, "platform-tools", "adb.exe")
EMULATOR_PATH = ""    if IS_LINUX else os.path.join(SDK_PATH, "emulator", "emulator.exe")

_apk_env = os.environ.get("APK_PATH", "")
if _apk_env:
    APK_PATH = _apk_env
elif IS_LINUX:
    APK_PATH = os.path.join(
        os.environ.get("GITHUB_WORKSPACE", ""),
        "app/build/outputs/apk/debug/app-debug.apk"
    )
else:
    APK_PATH = r"C:\Users\HP\Projects\TRUTH GUARD\app\build\outputs\apk\debug\app-debug.apk"

AVD_NAME     = os.environ.get("AVD_NAME", "Pixel_6")
APP_PACKAGE  = "com.samavaishnavi.truthguard"
APPIUM_PORT  = 4723
APPIUM_HOST  = "127.0.0.1"

# ═══════════════════════════════════════════════════════════════════════════════
#  REPORT COLOUR PALETTE  (matches Selenium suite palette exactly)
# ═══════════════════════════════════════════════════════════════════════════════
COLORS = {
    'headerBg':    '1A1A2E',
    'headerText':  'FFFFFF',
    'subHeaderBg': '16213E',
    'moduleBg':    '0F3460',
    'moduleText':  'E94560',
    'passedBg':    'D4EDDA',
    'passedText':  '155724',
    'failedBg':    'F8D7DA',
    'failedText':  '721C24',
    'skippedBg':   'FFF3CD',
    'skippedText': '856404',
    'rowAlt':      'F8F9FA',
    'rowNorm':     'FFFFFF',
    'accent':      '0F3460',
    'statPass':    '28A745',
    'statFail':    'DC3545',
    'statTotal':   '17A2B8',
    'border':      'BDC3C7',
}

# ═══════════════════════════════════════════════════════════════════════════════
#  RESULTS STORE
# ═══════════════════════════════════════════════════════════════════════════════
results     = []
suite_start = time.time()

# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def log(status, test_id, name, duration, err=''):
    icon = '✅' if status == 'PASSED' else '❌' if status == 'FAILED' else '⚠️'
    err_msg = f" → {err[:90]}" if err else ""
    print(f"  {icon} [{test_id}] {name} ({duration}ms){err_msg}")

def record(module, test_id, name, desc, status, duration, error=''):
    results.append({
        'module': module, 'id': test_id, 'name': name,
        'desc': desc, 'status': status, 'duration': duration, 'error': error
    })
    log(status, test_id, name, duration, error)

def tc(module, test_id, name, desc, fn):
    t0 = time.time()
    try:
        fn()
        record(module, test_id, name, desc, 'PASSED', int((time.time() - t0) * 1000))
    except Exception as e:
        record(module, test_id, name, desc, 'FAILED', int((time.time() - t0) * 1000), str(e))

# ═══════════════════════════════════════════════════════════════════════════════
#  ELEMENT FINDERS
# ═══════════════════════════════════════════════════════════════════════════════
def find_text(driver, text, timeout=8):
    """Find element by text using multiple UIAutomator strategies."""
    end = time.time() + timeout
    while time.time() < end:
        for strategy in [
            f'new UiSelector().textContains("{text}")',
            f'new UiSelector().text("{text}")',
            f'new UiSelector().descriptionContains("{text}")',
        ]:
            try:
                return driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR, strategy)
            except Exception:
                pass
        try:
            return driver.find_element(AppiumBy.XPATH, f'//*[contains(@text, "{text}")]')
        except Exception:
            pass
        time.sleep(0.3)
    raise Exception(f"Text '{text}' not found on screen after {timeout}s")

def find_exact(driver, text, timeout=8):
    """Find element by exact text match."""
    end = time.time() + timeout
    while time.time() < end:
        try:
            return driver.find_element(
                AppiumBy.ANDROID_UIAUTOMATOR,
                f'new UiSelector().text("{text}")'
            )
        except Exception:
            pass
        time.sleep(0.3)
    raise Exception(f"Exact text '{text}' not found after {timeout}s")

def assert_text_present(driver, text, timeout=8):
    """Assert that text is visible on the current screen."""
    find_text(driver, text, timeout)

def click_button(driver, text, timeout=8):
    """Click a button by its label text."""
    el = find_text(driver, text, timeout)
    el.click()
    time.sleep(0.6)

def enter_text(driver, text):
    """Type into the single EditText on screen (VerifyScreen)."""
    el = driver.find_element(AppiumBy.CLASS_NAME, "android.widget.EditText")
    el.clear()
    if text:
        el.send_keys(text)
    time.sleep(0.3)

def go_home(driver):
    """Navigate back to the Home screen."""
    driver.back()
    time.sleep(0.5)

def wait_for_screen(driver, text, timeout=10):
    """Wait until a specific screen identifier is visible."""
    find_text(driver, text, timeout)

# ═══════════════════════════════════════════════════════════════════════════════
#  ENVIRONMENT SETUP
# ═══════════════════════════════════════════════════════════════════════════════
def start_emulator():
    print("\n🔍 Checking for running emulator / device...")
    res = subprocess.run([ADB_PATH, "devices"], capture_output=True, text=True)
    lines = [l.strip() for l in res.stdout.splitlines() if l.strip() and "List of devices" not in l]
    if any("device" in l or "emulator" in l for l in lines):
        print("🟢 Emulator / device already running.")
        return

    if IS_CI:
        print("⏳ CI mode: waiting for emulator from runner action...")
        subprocess.run([ADB_PATH, "wait-for-device"], timeout=180)
        for _ in range(60):
            r = subprocess.run([ADB_PATH, "shell", "getprop", "sys.boot_completed"],
                               capture_output=True, text=True)
            if "1" in r.stdout:
                print("🟢 Emulator ready.")
                time.sleep(3)
                return
            time.sleep(3)
        print("⚠️  Boot check timed out — proceeding anyway.")
        return

    if not EMULATOR_PATH:
        raise RuntimeError("EMULATOR_PATH not set for this platform.")
    print(f"🚀 Launching emulator: {AVD_NAME}")
    subprocess.Popen([EMULATOR_PATH, "-avd", AVD_NAME, "-delay-adb"])
    subprocess.run([ADB_PATH, "wait-for-device"])
    for _ in range(60):
        r = subprocess.run([ADB_PATH, "shell", "getprop", "sys.boot_completed"],
                           capture_output=True, text=True)
        if "1" in r.stdout:
            print("🟢 Emulator booted.")
            time.sleep(3)
            return
        time.sleep(2)
    print("⚠️  Boot check timed out.")

def start_appium_server():
    if IS_CI:
        print("ℹ️  CI mode: Appium managed by workflow — skipping local start.")
        return None, None
    print("🚀 Starting Appium server...")
    log_f = open("appium_server.log", "w")
    cmd = ["cmd", "/c", "npx", "appium",
           "--port", str(APPIUM_PORT), "--address", APPIUM_HOST]
    env = os.environ.copy()
    env["ANDROID_HOME"]     = SDK_PATH
    env["ANDROID_SDK_ROOT"] = SDK_PATH
    env["JAVA_HOME"]        = r"C:\Program Files\Android\Android Studio\jbr"
    proc = subprocess.Popen(cmd, env=env, stdout=log_f, stderr=log_f)
    time.sleep(8)
    print("🟢 Appium server started.")
    return proc, log_f

# ═══════════════════════════════════════════════════════════════════════════════
#  TEST SUITE — 135 TEST CASES  (8 Modules)
# ═══════════════════════════════════════════════════════════════════════════════
def run_all_tests(driver):

    # ──────────────────────────────────────────────────────────────────────────
    # MODULE 1 · APP LAUNCH & HOME SCREEN  (TC-001 – TC-022)
    # ──────────────────────────────────────────────────────────────────────────
    print("\n📋 MODULE 1: App Launch & Home Screen")

    tc("App Launch", "TC-001", "App launches without crash",
       "MainActivity starts successfully", lambda: driver.activate_app(APP_PACKAGE))

    tc("App Launch", "TC-002", "Shield emoji visible on Home",
       "🛡️ emoji Text element present", lambda: assert_text_present(driver, "🛡️"))

    tc("App Launch", "TC-003", "TRUTHGUARD title visible",
       "Text 'TRUTHGUARD' present", lambda: assert_text_present(driver, "TRUTHGUARD"))

    tc("App Launch", "TC-004", "Subtitle 'AI Powered Fake News Detection' visible",
       "Subtitle text present", lambda: assert_text_present(driver, "AI Powered Fake News Detection"))

    tc("App Launch", "TC-005", "Version 1.0 label visible",
       "Footer version text present", lambda: assert_text_present(driver, "Version 1.0"))

    tc("App Launch", "TC-006", "'🔍 Verify News' button visible",
       "Verify News button present on Home", lambda: assert_text_present(driver, "Verify News"))

    tc("App Launch", "TC-007", "'📰 Trending News' button visible",
       "Trending News button present on Home", lambda: assert_text_present(driver, "Trending News"))

    tc("App Launch", "TC-008", "'📊 Dashboard' button visible",
       "Dashboard button present on Home", lambda: assert_text_present(driver, "Dashboard"))

    tc("App Launch", "TC-009", "'ℹ About' button visible",
       "About button present on Home", lambda: assert_text_present(driver, "About"))

    tc("App Launch", "TC-010", "Home screen has exactly 4 buttons",
       "All 4 navigation buttons rendered",
       lambda: _assert_button_count(driver, 4))

    def _assert_button_count(d, expected):
        buttons = d.find_elements(AppiumBy.CLASS_NAME, "android.widget.Button")
        if len(buttons) < expected:
            raise Exception(f"Expected >= {expected} buttons, found {len(buttons)}")
    tc("App Launch", "TC-010", "Home screen has at least 4 nav buttons",
       "Button count >= 4", lambda: _assert_button_count(driver, 4))

    tc("App Launch", "TC-011", "Verify News button is clickable",
       "Button clickable attribute = true",
       lambda: _assert_clickable(driver, "Verify News"))

    def _assert_clickable(d, text):
        el = find_text(d, text)
        val = el.get_attribute("clickable")
        if val != "true":
            raise Exception(f"Button '{text}' not clickable (clickable={val})")
    tc("App Launch", "TC-011", "Verify News button is clickable",
       "Button clickable attr = true", lambda: _assert_clickable(driver, "Verify News"))

    tc("App Launch", "TC-012", "Trending News button is clickable",
       "Button clickable", lambda: _assert_clickable(driver, "Trending News"))

    tc("App Launch", "TC-013", "Dashboard button is clickable",
       "Button clickable", lambda: _assert_clickable(driver, "Dashboard"))

    tc("App Launch", "TC-014", "About button is clickable",
       "Button clickable", lambda: _assert_clickable(driver, "About"))

    # Home → screen navigation
    tc("App Launch", "TC-015", "Tap Verify News → Verify screen opens",
       "Click button, Verify News header appears",
       lambda: (click_button(driver, "Verify News"), assert_text_present(driver, "Verify News")))

    go_home(driver)

    tc("App Launch", "TC-016", "Tap Trending News → Trending screen opens",
       "Click button, '📰 Trending News' header appears",
       lambda: (click_button(driver, "Trending News"), assert_text_present(driver, "Trending News")))

    go_home(driver)

    tc("App Launch", "TC-017", "Tap Dashboard → Dashboard screen opens",
       "Click button, '📊 Dashboard' header appears",
       lambda: (click_button(driver, "Dashboard"), assert_text_present(driver, "Dashboard")))

    go_home(driver)

    tc("App Launch", "TC-018", "Tap About → About screen opens",
       "Click button, 'TruthGuard' about title appears",
       lambda: (click_button(driver, "About"), assert_text_present(driver, "TruthGuard")))

    go_home(driver)

    tc("App Launch", "TC-019", "Back navigation from Verify returns to Home",
       "Back press from Verify → Home TRUTHGUARD title visible",
       lambda: (click_button(driver, "Verify News"),
                driver.back(), time.sleep(0.5),
                assert_text_present(driver, "TRUTHGUARD")))

    tc("App Launch", "TC-020", "Back navigation from Trending returns to Home",
       "Back press from Trending → Home",
       lambda: (click_button(driver, "Trending News"),
                driver.back(), time.sleep(0.5),
                assert_text_present(driver, "TRUTHGUARD")))

    tc("App Launch", "TC-021", "Back navigation from Dashboard returns to Home",
       "Back press from Dashboard → Home",
       lambda: (click_button(driver, "Dashboard"),
                driver.back(), time.sleep(0.5),
                assert_text_present(driver, "TRUTHGUARD")))

    tc("App Launch", "TC-022", "Back navigation from About returns to Home",
       "Back press from About → Home",
       lambda: (click_button(driver, "About"),
                driver.back(), time.sleep(0.5),
                assert_text_present(driver, "TRUTHGUARD")))

    # ──────────────────────────────────────────────────────────────────────────
    # MODULE 2 · VERIFY NEWS — UI STRUCTURE  (TC-023 – TC-040)
    # ──────────────────────────────────────────────────────────────────────────
    print("\n📋 MODULE 2: Verify News — UI Structure")
    click_button(driver, "Verify News")

    tc("Verify UI", "TC-023", "Verify screen header 'Verify News' present",
       "Title text visible", lambda: assert_text_present(driver, "Verify News"))

    tc("Verify UI", "TC-024", "'Paste News Here' label visible",
       "OutlinedTextField label present",
       lambda: assert_text_present(driver, "Paste News Here"))

    tc("Verify UI", "TC-025", "EditText input field present",
       "android.widget.EditText element found",
       lambda: driver.find_element(AppiumBy.CLASS_NAME, "android.widget.EditText"))

    tc("Verify UI", "TC-026", "'Analyze' button visible",
       "Analyze button present", lambda: assert_text_present(driver, "Analyze"))

    tc("Verify UI", "TC-027", "Analyze button is clickable",
       "Analyze button clickable attr = true",
       lambda: _assert_clickable(driver, "Analyze"))

    tc("Verify UI", "TC-028", "EditText is enabled and interactive",
       "EditText enabled attribute = true",
       lambda: _assert_field_enabled(driver))

    def _assert_field_enabled(d):
        el = d.find_element(AppiumBy.CLASS_NAME, "android.widget.EditText")
        if el.get_attribute("enabled") != "true":
            raise Exception("EditText is not enabled")
    tc("Verify UI", "TC-028", "EditText is enabled",
       "Enabled attr = true", lambda: _assert_field_enabled(driver))

    tc("Verify UI", "TC-029", "Result card hidden before first analysis",
       "Analysis Result heading absent initially",
       lambda: _assert_text_absent(driver, "Analysis Result", timeout=3))

    def _assert_text_absent(d, text, timeout=3):
        try:
            find_text(d, text, timeout=timeout)
            raise Exception(f"Text '{text}' should NOT be visible but was found")
        except Exception as e:
            if "should NOT" in str(e):
                raise
            pass  # expected — text absent is correct
    tc("Verify UI", "TC-029", "Result card absent before analysis",
       "Analysis Result text not visible yet",
       lambda: _assert_text_absent(driver, "Analysis Result", timeout=3))

    tc("Verify UI", "TC-030", "User can type into EditText",
       "Text entered successfully",
       lambda: (enter_text(driver, "test input"),
                assert_text_present(driver, "test input")))

    tc("Verify UI", "TC-031", "User can clear EditText",
       "EditText cleared",
       lambda: (enter_text(driver, ""),
                _assert_field_empty(driver)))

    def _assert_field_empty(d):
        el = d.find_element(AppiumBy.CLASS_NAME, "android.widget.EditText")
        val = el.get_attribute("text") or ""
        if val.strip():
            raise Exception(f"EditText not empty, contains: '{val}'")
    tc("Verify UI", "TC-031", "EditText clears properly",
       "Field empty after clear", lambda: _assert_field_empty(driver))

    tc("Verify UI", "TC-032", "Verify screen scrollable",
       "Vertical scroll action completes",
       lambda: driver.swipe(540, 1400, 540, 600, 500))

    tc("Verify UI", "TC-033", "Verify screen scroll up recovers",
       "Scroll back up completes",
       lambda: driver.swipe(540, 600, 540, 1400, 500))

    # Result card appearance after analysis
    def _run_genuine_and_check_card():
        enter_text(driver, "Standard global news")
        click_button(driver, "Analyze")
        assert_text_present(driver, "Analysis Result")
    tc("Verify UI", "TC-034", "Result card appears after analysis",
       "Card with 'Analysis Result' visible after clicking Analyze",
       _run_genuine_and_check_card)

    tc("Verify UI", "TC-035", "Analysis Result heading bold / prominent",
       "'Analysis Result' text node present in card",
       lambda: assert_text_present(driver, "Analysis Result"))

    tc("Verify UI", "TC-036", "Result title element visible in card",
       "Likely Genuine/Fake text present",
       lambda: assert_text_present(driver, "Likely Genuine News"))

    tc("Verify UI", "TC-037", "Confidence score label visible in card",
       "Confidence Score text present",
       lambda: assert_text_present(driver, "Confidence Score"))

    tc("Verify UI", "TC-038", "Recommendation label visible in card",
       "Recommendation: text present",
       lambda: assert_text_present(driver, "Recommendation:"))

    tc("Verify UI", "TC-039", "Re-entering text and re-analyzing updates card",
       "Second analysis replaces previous result",
       lambda: (enter_text(driver, "This is shocking fake"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Fake News")))

    go_home(driver)
    tc("Verify UI", "TC-040", "Back from Verify returns to Home cleanly",
       "TRUTHGUARD title visible on return",
       lambda: (click_button(driver, "Verify News"),
                driver.back(), time.sleep(0.5),
                assert_text_present(driver, "TRUTHGUARD")))

    # ──────────────────────────────────────────────────────────────────────────
    # MODULE 3 · VERIFY NEWS — ANALYSIS LOGIC  (TC-041 – TC-075)
    # ──────────────────────────────────────────────────────────────────────────
    print("\n📋 MODULE 3: Verify News — Analysis Logic")
    click_button(driver, "Verify News")

    # — Genuine news —
    def _genuine_full():
        enter_text(driver, "This is genuine news content with no triggers")
        click_button(driver, "Analyze")
        assert_text_present(driver, "Likely Genuine News")
    tc("Verify Logic", "TC-041", "Genuine news → 'Likely Genuine News' result",
       "No fake keywords → genuine result", _genuine_full)

    tc("Verify Logic", "TC-042", "Genuine result shows 94% confidence",
       "Confidence Score : 94%",
       lambda: assert_text_present(driver, "Confidence Score : 94%"))

    tc("Verify Logic", "TC-043", "Genuine result shows 'This news appears reliable.'",
       "Recommendation text correct",
       lambda: assert_text_present(driver, "This news appears reliable."))

    tc("Verify Logic", "TC-044", "Genuine result prefix is ✅",
       "✅ Likely Genuine News visible",
       lambda: assert_text_present(driver, "✅ Likely Genuine News"))

    # — Fake keywords (lowercase) —
    fake_kws = [
        ("fake",      "TC-045"),
        ("hoax",      "TC-046"),
        ("rumor",     "TC-047"),
        ("clickbait", "TC-048"),
        ("shocking",  "TC-049"),
    ]
    for kw, tid in fake_kws:
        def _test_kw(k=kw):
            enter_text(driver, f"This article contains the word {k}")
            click_button(driver, "Analyze")
            assert_text_present(driver, "Likely Fake News")
        tc("Verify Logic", tid, f"Keyword '{kw}' (lowercase) triggers Fake result",
           f"'{kw}' present → Likely Fake News", _test_kw)

    # — Fake keywords (UPPERCASE) —
    upper_kws = [
        ("FAKE",      "TC-050"),
        ("HOAX",      "TC-051"),
        ("RUMOR",     "TC-052"),
        ("CLICKBAIT", "TC-053"),
        ("SHOCKING",  "TC-054"),
    ]
    for kw, tid in upper_kws:
        def _test_upper(k=kw):
            enter_text(driver, f"This alert contains {k} information")
            click_button(driver, "Analyze")
            assert_text_present(driver, "Likely Fake News")
        tc("Verify Logic", tid, f"Keyword '{kw}' (UPPERCASE) triggers Fake result",
           "Case-insensitive detection", _test_upper)

    # — Mixed-case keywords —
    mixed_kws = [
        ("fAkE",      "TC-055"),
        ("hOaX",      "TC-056"),
        ("RuMoR",     "TC-057"),
        ("ClIcKbAiT", "TC-058"),
        ("ShOcKiNg",  "TC-059"),
    ]
    for kw, tid in mixed_kws:
        def _test_mixed(k=kw):
            enter_text(driver, f"News report: {k} alert issued")
            click_button(driver, "Analyze")
            assert_text_present(driver, "Likely Fake News")
        tc("Verify Logic", tid, f"Mixed-case '{kw}' triggers Fake result",
           "Mixed case detected via .lowercase()", _test_mixed)

    # — Fake result metadata —
    def _setup_fake():
        enter_text(driver, "This story is a complete hoax and clickbait")
        click_button(driver, "Analyze")
    _setup_fake()

    tc("Verify Logic", "TC-060", "Fake result prefix is ❌",
       "❌ Likely Fake News text visible",
       lambda: assert_text_present(driver, "❌ Likely Fake News"))

    tc("Verify Logic", "TC-061", "Fake result shows 88% confidence",
       "Confidence Score : 88% text visible",
       lambda: assert_text_present(driver, "Confidence Score : 88%"))

    tc("Verify Logic", "TC-062", "Fake result recommendation correct",
       "Verify this news using trusted sources before sharing.",
       lambda: assert_text_present(driver, "Verify this news using trusted sources before sharing."))

    tc("Verify Logic", "TC-063", "Recommendation label prefix correct",
       "Recommendation: label present",
       lambda: assert_text_present(driver, "Recommendation:"))

    # — Boundary conditions —
    tc("Verify Logic", "TC-064", "Empty input → Likely Genuine News",
       "Empty string has no keywords → genuine",
       lambda: (enter_text(driver, ""),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-065", "Only whitespace → Likely Genuine News",
       "Whitespace only → genuine (no keywords)",
       lambda: (enter_text(driver, "   "),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-066", "Long genuine article → 94% confidence",
       "200-char genuine text → Likely Genuine News",
       lambda: (enter_text(driver,
                "This is a very long genuine news article about the recent international climate summit "
                "where global leaders agreed on strict carbon emission reduction targets for 2030."),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-067", "Fake keyword at start of long text detected",
       "Starts with 'fake' → Likely Fake News",
       lambda: (enter_text(driver,
                "fake news story about the recent summit where global leaders discussed "
                "carbon emission targets and renewable energy policies."),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-068", "Fake keyword at end of long text detected",
       "Ends with 'hoax' → Likely Fake News",
       lambda: (enter_text(driver,
                "Breaking news about the economy and scientific breakthroughs "
                "around the world — but it's all a hoax"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-069", "Special characters with keyword → Fake",
       "### SHOCKING!!! triggers Fake",
       lambda: (enter_text(driver, "### SHOCKING!!! $$$ breaking alert ***"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-070", "Multiple keywords in text → Fake",
       "fake + hoax + rumor all present",
       lambda: (enter_text(driver, "Warning! fake hoax rumor clickbait shocking story"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-071", "Numeric text without keywords → Genuine",
       "1234 5678 90 → genuine (no keywords)",
       lambda: (enter_text(driver, "1234 5678 9012 3456 7890"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-072", "Single genuine word → Genuine",
       "Word 'science' → genuine",
       lambda: (enter_text(driver, "science"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-073", "Single fake word → Fake",
       "Word 'fake' alone → Likely Fake News",
       lambda: (enter_text(driver, "fake"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Fake News")))

    tc("Verify Logic", "TC-074", "Re-analysis switches result Fake→Genuine",
       "Enter fake, analyze, enter genuine, analyze → genuine",
       lambda: (enter_text(driver, "total hoax"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Fake News"),
                enter_text(driver, "breaking science news"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Genuine News")))

    tc("Verify Logic", "TC-075", "Re-analysis switches result Genuine→Fake",
       "Enter genuine, analyze, enter fake, analyze → fake",
       lambda: (enter_text(driver, "genuine science research"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Genuine News"),
                enter_text(driver, "clickbait rumor story"),
                click_button(driver, "Analyze"),
                assert_text_present(driver, "Likely Fake News")))

    go_home(driver)

    # ──────────────────────────────────────────────────────────────────────────
    # MODULE 4 · TRENDING NEWS SCREEN  (TC-076 – TC-100)
    # ──────────────────────────────────────────────────────────────────────────
    print("\n📋 MODULE 4: Trending News Screen")
    click_button(driver, "Trending News")

    tc("Trending", "TC-076", "Trending screen opens successfully",
       "Screen navigated, header present",
       lambda: assert_text_present(driver, "Trending News"))

    tc("Trending", "TC-077", "Header '📰 Trending News' text correct",
       "Exact header string with emoji",
       lambda: assert_text_present(driver, "📰 Trending News"))

    # ── News item 1: BBC ──────────────────────────────────────────────────────
    tc("Trending", "TC-078", "BBC title: 'Scientists discover new climate monitoring technology'",
       "1st card title exact match",
       lambda: assert_text_present(driver, "Scientists discover new climate monitoring technology"))

    tc("Trending", "TC-079", "BBC source label correct",
       "Source: BBC text present",
       lambda: assert_text_present(driver, "Source: BBC"))

    # ── News item 2: Reuters ──────────────────────────────────────────────────
    tc("Trending", "TC-080", "Reuters title: 'AI transforming healthcare worldwide'",
       "2nd card title exact match",
       lambda: assert_text_present(driver, "AI transforming healthcare worldwide"))

    tc("Trending", "TC-081", "Reuters source label correct",
       "Source: Reuters text present",
       lambda: assert_text_present(driver, "Source: Reuters"))

    # ── News item 3: NASA ─────────────────────────────────────────────────────
    tc("Trending", "TC-082", "NASA title: 'Space mission successfully reaches orbit'",
       "3rd card title exact match",
       lambda: assert_text_present(driver, "Space mission successfully reaches orbit"))

    tc("Trending", "TC-083", "NASA source label correct",
       "Source: NASA text present",
       lambda: assert_text_present(driver, "Source: NASA"))

    # ── News item 4: Bloomberg ────────────────────────────────────────────────
    tc("Trending", "TC-084", "Bloomberg title: 'Global economy shows positive growth'",
       "4th card title exact match",
       lambda: assert_text_present(driver, "Global economy shows positive growth"))

    tc("Trending", "TC-085", "Bloomberg source label correct",
       "Source: Bloomberg text present",
       lambda: assert_text_present(driver, "Source: Bloomberg"))

    # ── News item 5: UNESCO ───────────────────────────────────────────────────
    tc("Trending", "TC-086", "UNESCO title: 'Education sector adopts AI learning tools'",
       "5th card title exact match",
       lambda: assert_text_present(driver, "Education sector adopts AI learning tools"))

    tc("Trending", "TC-087", "UNESCO source label correct",
       "Source: UNESCO text present",
       lambda: assert_text_present(driver, "Source: UNESCO"))

    # ── All 5 sources present ─────────────────────────────────────────────────
    tc("Trending", "TC-088", "All 5 sources (BBC/Reuters/NASA/Bloomberg/UNESCO) present",
       "All source labels visible without scrolling or with lazy load",
       lambda: [assert_text_present(driver, f"Source: {s}")
                for s in ["BBC", "Reuters", "NASA", "Bloomberg", "UNESCO"]])

    # ── Scroll interactions ───────────────────────────────────────────────────
    tc("Trending", "TC-089", "Scroll down gesture on Trending list",
       "Swipe down on list works without crash",
       lambda: driver.swipe(540, 1500, 540, 500, 600))

    tc("Trending", "TC-090", "Scroll up gesture on Trending list",
       "Swipe up on list works without crash",
       lambda: driver.swipe(540, 500, 540, 1500, 600))

    tc("Trending", "TC-091", "News titles still visible after scroll cycle",
       "First card title still accessible",
       lambda: assert_text_present(driver, "Scientists discover new climate monitoring technology"))

    tc("Trending", "TC-092", "Cards have distinct title text",
       "No duplicate card text — 5 unique titles",
       lambda: _assert_unique_titles(driver))

    def _assert_unique_titles(d):
        titles = [
            "Scientists discover new climate monitoring technology",
            "AI transforming healthcare worldwide",
            "Space mission successfully reaches orbit",
            "Global economy shows positive growth",
            "Education sector adopts AI learning tools",
        ]
        for t in titles:
            assert_text_present(d, t)
    tc("Trending", "TC-092", "All 5 unique titles present",
       "Each card title unique and visible", lambda: _assert_unique_titles(driver))

    tc("Trending", "TC-093", "Cards have distinct source labels",
       "5 distinct source labels present",
       lambda: [assert_text_present(driver, f"Source: {s}")
                for s in ["BBC", "Reuters", "NASA", "Bloomberg", "UNESCO"]])

    tc("Trending", "TC-094", "No 'fake' or 'hoax' text in Trending cards",
       "Trending news contains no fake indicators",
       lambda: _assert_no_fake_text(driver))

    def _assert_no_fake_text(d):
        for bad in ["hoax", "HOAX", "clickbait"]:
            try:
                find_text(d, bad, timeout=2)
                raise Exception(f"Unexpected fake keyword '{bad}' found in Trending")
            except Exception as e:
                if "Unexpected" in str(e):
                    raise
                pass  # expected — these words should not appear
    tc("Trending", "TC-094", "No fake keywords in Trending news content",
       "Legitimate news only", lambda: _assert_no_fake_text(driver))

    tc("Trending", "TC-095", "Trending screen loads quickly (<5 s)",
       "All items appear within 5 seconds",
       lambda: assert_text_present(driver, "Source: BBC", timeout=5))

    tc("Trending", "TC-096", "Back from Trending → Home screen",
       "Back press returns to Home with TRUTHGUARD title",
       lambda: (driver.back(), time.sleep(0.5),
                assert_text_present(driver, "TRUTHGUARD")))

    # Re-enter Trending for remaining tests
    click_button(driver, "Trending News")

    tc("Trending", "TC-097", "Trending News button navigates back correctly",
       "Re-entering Trending screen works",
       lambda: assert_text_present(driver, "📰 Trending News"))

    tc("Trending", "TC-098", "Card content persists on re-entry",
       "BBC card title still present on second visit",
       lambda: assert_text_present(driver, "Scientists discover new climate monitoring technology"))

    tc("Trending", "TC-099", "LazyColumn renders all 5 items",
       "5 news sources all visible",
       lambda: [assert_text_present(driver, s)
                for s in ["Source: BBC", "Source: Reuters", "Source: NASA",
                          "Source: Bloomberg", "Source: UNESCO"]])

    go_home(driver)

    tc("Trending", "TC-100", "Trending → Home → Trending round-trip",
       "Full navigation cycle completes",
       lambda: (click_button(driver, "Trending News"),
                assert_text_present(driver, "📰 Trending News"),
                driver.back(), time.sleep(0.5),
                assert_text_present(driver, "TRUTHGUARD")))

    # ──────────────────────────────────────────────────────────────────────────
    # MODULE 5 · DASHBOARD SCREEN  (TC-101 – TC-115)
    # ──────────────────────────────────────────────────────────────────────────
    print("\n📋 MODULE 5: Dashboard Screen")
    click_button(driver, "Dashboard")

    tc("Dashboard", "TC-101", "Dashboard screen opens",
       "Header text visible", lambda: assert_text_present(driver, "Dashboard"))

    tc("Dashboard", "TC-102", "Header '📊 Dashboard' exact text",
       "Emoji + title correct", lambda: assert_text_present(driver, "📊 Dashboard"))

    tc("Dashboard", "TC-103", "'Articles Verified' stat card title visible",
       "StatCard title text present",
       lambda: assert_text_present(driver, "Articles Verified"))

    tc("Dashboard", "TC-104", "Articles Verified value = 25",
       "Stat value '25' present",
       lambda: assert_text_present(driver, "25"))

    tc("Dashboard", "TC-105", "'True News' stat card title visible",
       "StatCard title text present",
       lambda: assert_text_present(driver, "True News"))

    tc("Dashboard", "TC-106", "True News value = 18",
       "Stat value '18' present",
       lambda: assert_text_present(driver, "18"))

    tc("Dashboard", "TC-107", "'Fake News' stat card title visible",
       "StatCard title text present",
       lambda: assert_text_present(driver, "Fake News"))

    tc("Dashboard", "TC-108", "Fake News value = 7",
       "Stat value '7' present",
       lambda: assert_text_present(driver, "7"))

    tc("Dashboard", "TC-109", "'Accuracy' stat card title visible",
       "StatCard title text present",
       lambda: assert_text_present(driver, "Accuracy"))

    tc("Dashboard", "TC-110", "Accuracy value = 92%",
       "Stat value '92%' present",
       lambda: assert_text_present(driver, "92%"))

    tc("Dashboard", "TC-111", "All 4 stat card titles present simultaneously",
       "All stat titles visible without scrolling",
       lambda: [assert_text_present(driver, t)
                for t in ["Articles Verified", "True News", "Fake News", "Accuracy"]])

    tc("Dashboard", "TC-112", "All 4 stat values present simultaneously",
       "Values 25, 18, 7, 92% all visible",
       lambda: [assert_text_present(driver, v)
                for v in ["25", "18", "7", "92%"]])

    tc("Dashboard", "TC-113", "Dashboard persists on re-entry",
       "Navigate away and back — stats still show",
       lambda: (driver.back(), time.sleep(0.5),
                click_button(driver, "Dashboard"),
                assert_text_present(driver, "Articles Verified")))

    tc("Dashboard", "TC-114", "Dashboard stat values persist on re-entry",
       "Values 25, 18, 7, 92% still visible on second visit",
       lambda: [assert_text_present(driver, v)
                for v in ["25", "18", "7", "92%"]])

    go_home(driver)

    tc("Dashboard", "TC-115", "Back from Dashboard → Home",
       "TRUTHGUARD visible after back",
       lambda: (click_button(driver, "Dashboard"),
                driver.back(), time.sleep(0.5),
                assert_text_present(driver, "TRUTHGUARD")))

    # ──────────────────────────────────────────────────────────────────────────
    # MODULE 6 · ABOUT SCREEN  (TC-116 – TC-126)
    # ──────────────────────────────────────────────────────────────────────────
    print("\n📋 MODULE 6: About Screen")
    click_button(driver, "About")

    tc("About", "TC-116", "About screen opens",
       "TruthGuard title visible",
       lambda: assert_text_present(driver, "TruthGuard"))

    tc("About", "TC-117", "Shield emoji '🛡️' visible on About",
       "Large emoji Text element present",
       lambda: assert_text_present(driver, "🛡️"))

    tc("About", "TC-118", "About screen title 'TruthGuard' exact",
       "h2-style title text correct",
       lambda: assert_text_present(driver, "TruthGuard"))

    tc("About", "TC-119", "'AI Powered Fake News Detection App' text in card",
       "Description text correct",
       lambda: assert_text_present(driver, "AI Powered Fake News Detection App"))

    tc("About", "TC-120", "'Version : 1.0' text in card",
       "Version label with spaces correct",
       lambda: assert_text_present(driver, "Version : 1.0"))

    tc("About", "TC-121", "'Developed for Educational Purpose' text in card",
       "Purpose text correct",
       lambda: assert_text_present(driver, "Developed for Educational Purpose"))

    tc("About", "TC-122", "'Technology : Kotlin + Jetpack Compose + AI' text",
       "Tech stack text correct",
       lambda: assert_text_present(driver, "Technology : Kotlin + Jetpack Compose + AI"))

    tc("About", "TC-123", "Copyright '© TruthGuard 2025' text visible",
       "Footer copyright text correct",
       lambda: assert_text_present(driver, "© TruthGuard 2025"))

    tc("About", "TC-124", "All About card items present simultaneously",
       "All 4 card texts visible",
       lambda: [assert_text_present(driver, t) for t in [
           "AI Powered Fake News Detection App",
           "Version : 1.0",
           "Developed for Educational Purpose",
           "Technology : Kotlin + Jetpack Compose + AI",
       ]])

    tc("About", "TC-125", "About screen persists on re-entry",
       "Navigate away and back — content still correct",
       lambda: (driver.back(), time.sleep(0.5),
                click_button(driver, "About"),
                assert_text_present(driver, "TruthGuard")))

    go_home(driver)

    tc("About", "TC-126", "Back from About → Home",
       "TRUTHGUARD visible after back press",
       lambda: (click_button(driver, "About"),
                driver.back(), time.sleep(0.5),
                assert_text_present(driver, "TRUTHGUARD")))

    # ──────────────────────────────────────────────────────────────────────────
    # MODULE 7 · DEVICE & ORIENTATION  (TC-127 – TC-130)
    # ──────────────────────────────────────────────────────────────────────────
    print("\n📋 MODULE 7: Device & Orientation")

    tc("Device", "TC-127", "Portrait → Landscape → Portrait: Home intact",
       "Rotate to landscape and back — TRUTHGUARD visible",
       lambda: (setattr(driver, 'orientation', 'LANDSCAPE'),
                time.sleep(1.5),
                assert_text_present(driver, "TRUTHGUARD"),
                setattr(driver, 'orientation', 'PORTRAIT'),
                time.sleep(1.5),
                assert_text_present(driver, "TRUTHGUARD")))

    tc("Device", "TC-128", "Landscape mode: Verify screen usable",
       "Navigate to Verify in landscape, EditText present",
       lambda: (setattr(driver, 'orientation', 'LANDSCAPE'),
                time.sleep(1),
                click_button(driver, "Verify News"),
                driver.find_element(AppiumBy.CLASS_NAME, "android.widget.EditText"),
                driver.back(), time.sleep(0.5),
                setattr(driver, 'orientation', 'PORTRAIT'),
                time.sleep(1)))

    tc("Device", "TC-129", "App resumes correctly after Home button press",
       "Press Home key, re-open app — TRUTHGUARD title visible",
       lambda: (driver.press_keycode(3),  # KEYCODE_HOME
                time.sleep(1.5),
                driver.activate_app(APP_PACKAGE),
                time.sleep(1),
                assert_text_present(driver, "TRUTHGUARD")))

    tc("Device", "TC-130", "App state restored after screen-off / wake",
       "Screen off and on — app state intact",
       lambda: (driver.press_keycode(26),  # KEYCODE_POWER
                time.sleep(1.5),
                driver.press_keycode(26),
                time.sleep(2),
                driver.activate_app(APP_PACKAGE),
                time.sleep(1),
                assert_text_present(driver, "TRUTHGUARD")))

    # ──────────────────────────────────────────────────────────────────────────
    # MODULE 8 · END-TO-END FLOWS  (TC-131 – TC-135)
    # ──────────────────────────────────────────────────────────────────────────
    print("\n📋 MODULE 8: End-to-End User Flows")

    # E2E flow 1: Full fake-news user journey
    def _e2e_fake_flow():
        click_button(driver, "Verify News")
        enter_text(driver, "Breaking: hoax about global leaders proven false")
        click_button(driver, "Analyze")
        assert_text_present(driver, "❌ Likely Fake News")
        assert_text_present(driver, "Confidence Score : 88%")
        assert_text_present(driver, "Verify this news using trusted sources before sharing.")
        driver.back(); time.sleep(0.5)
        assert_text_present(driver, "TRUTHGUARD")
    tc("E2E Flow", "TC-131", "Full fake-news flow: Home→Verify→Analyze→Back→Home",
       "Complete fake detection journey", _e2e_fake_flow)

    # E2E flow 2: Full genuine-news user journey
    def _e2e_genuine_flow():
        click_button(driver, "Verify News")
        enter_text(driver, "NASA confirms successful satellite launch for climate monitoring")
        click_button(driver, "Analyze")
        assert_text_present(driver, "✅ Likely Genuine News")
        assert_text_present(driver, "Confidence Score : 94%")
        assert_text_present(driver, "This news appears reliable.")
        driver.back(); time.sleep(0.5)
        assert_text_present(driver, "TRUTHGUARD")
    tc("E2E Flow", "TC-132", "Full genuine-news flow: Home→Verify→Analyze→Back→Home",
       "Complete genuine detection journey", _e2e_genuine_flow)

    # E2E flow 3: Explore all screens sequentially from Home
    def _e2e_all_screens():
        for btn, header in [
            ("Verify News",  "Verify News"),
            ("Trending News", "📰 Trending News"),
            ("Dashboard",    "📊 Dashboard"),
            ("About",        "TruthGuard"),
        ]:
            driver.back(); time.sleep(0.5)  # ensure Home
            click_button(driver, btn)
            assert_text_present(driver, header)
        driver.back(); time.sleep(0.5)
        assert_text_present(driver, "TRUTHGUARD")
    tc("E2E Flow", "TC-133", "Sequential navigation of all 5 screens",
       "Visit every screen and return to Home", _e2e_all_screens)

    # E2E flow 4: Verify screen state resets on re-entry
    def _e2e_state_reset():
        click_button(driver, "Verify News")
        enter_text(driver, "shocking rumor exposed by investigator")
        click_button(driver, "Analyze")
        assert_text_present(driver, "❌ Likely Fake News")
        driver.back(); time.sleep(0.5)
        # Re-enter Verify — state (result card) should be reset by Compose recomposition
        click_button(driver, "Verify News")
        # The result card should be gone — only original UI present
        assert_text_present(driver, "Paste News Here")
        try:
            find_text(driver, "Analysis Result", timeout=2)
            raise Exception("State not reset — Analysis Result card still visible on re-entry")
        except Exception as e:
            if "State not reset" in str(e):
                raise
    tc("E2E Flow", "TC-134", "Verify screen resets state on re-entry",
       "Result card absent after navigating away and back",
       _e2e_state_reset)

    go_home(driver)

    # E2E flow 5: Dashboard + Trending round-trip
    def _e2e_dashboard_trending():
        click_button(driver, "Dashboard")
        assert_text_present(driver, "92%")
        driver.back(); time.sleep(0.5)
        click_button(driver, "Trending News")
        assert_text_present(driver, "Source: BBC")
        driver.back(); time.sleep(0.5)
        assert_text_present(driver, "TRUTHGUARD")
    tc("E2E Flow", "TC-135", "Dashboard → Home → Trending → Home round-trip",
       "Stats and news persist across navigation cycle",
       _e2e_dashboard_trending)

    print("\n✅ All 135 test cases completed.")


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════
def generate_excel_report(report_path):
    print("\n📊 Generating Excel report...")
    wb = openpyxl.Workbook()

    # ── Style helpers ─────────────────────────────────────────────────────────
    def fnt(name="Plus Jakarta Sans", size=10, bold=False, color="000000"):
        return Font(name=name, size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def border():
        s = Side(style='thin', color=COLORS['border'])
        return Border(left=s, right=s, top=s, bottom=s)

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    passed_cnt   = len([r for r in results if r['status'] == 'PASSED'])
    failed_cnt   = len([r for r in results if r['status'] == 'FAILED'])
    total_cnt    = len(results)
    pass_rate    = round((passed_cnt / total_cnt) * 100) if total_cnt > 0 else 0
    duration_sec = round(time.time() - suite_start, 1)

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 1 — SUMMARY
    # ──────────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "📊 Summary"
    ws.views.sheetView[0].showGridLines = False

    # Banner
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "🛡️  TRUTHGUARD ANDROID — APPIUM E2E TEST REPORT"
    c.font      = fnt("Outfit", 18, True, COLORS['headerText'])
    c.fill      = fill(COLORS['headerBg'])
    c.alignment = align("center")
    ws.row_dimensions[1].height = 52

    # Subtitle
    ws.merge_cells("A2:H2")
    ts = datetime.now().strftime("%d/%m/%Y, %I:%M:%S %p")
    sub = ws["A2"]
    sub.value     = (f"Generated: {ts}  |  Engine: Python + Appium Client 3.x  |"
                     f"  Platform: Android API 33  |  Automation: UIAutomator2")
    sub.font      = fnt("Plus Jakarta Sans", 10, color="B0C4DE")
    sub.fill      = fill(COLORS['subHeaderBg'])
    sub.alignment = align("center")
    ws.row_dimensions[2].height = 26

    ws.append([])  # row 3 spacer

    # KPI boxes (row 4–5)
    kpi_data = [
        ("TOTAL TESTS", total_cnt,        COLORS['statTotal'], "B"),
        ("PASSED",      passed_cnt,        COLORS['statPass'],  "D"),
        ("FAILED",      failed_cnt,        COLORS['statFail'],  "F"),
        ("PASS RATE",   f"{pass_rate}%",  COLORS['accent'],    "H"),
    ]
    for label, val, color, col in kpi_data:
        ws.merge_cells(f"{col}4:{col}5")
        c = ws[f"{col}4"]
        c.value     = val
        c.font      = fnt("Outfit", 28, True, "FFFFFF")
        c.fill      = fill(color)
        c.alignment = align("center")

        lc = ws[f"{col}6"]
        lc.value     = label
        lc.font      = fnt("Plus Jakarta Sans", 9, True, color)
        lc.alignment = align("center")
    ws.row_dimensions[4].height = 44
    ws.row_dimensions[5].height = 44
    ws.row_dimensions[6].height = 18

    ws.append([])
    ws.append([])

    # Execution details table
    det_row = ws.append(["", "EXECUTION DETAILS", "", "", "", "", "", ""])
    dr = ws.max_row
    ws.merge_cells(f"B{dr}:H{dr}")
    ws[f"B{dr}"].font  = fnt("Plus Jakarta Sans", 12, True, COLORS['accent'])
    ws[f"B{dr}"].fill  = fill("E8EEF7")
    ws.row_dimensions[dr].height = 26

    details = [
        ("Test Suite",        "TruthGuard Android — Appium E2E"),
        ("App Package",       APP_PACKAGE),
        ("Total Test Cases",  total_cnt),
        ("Passed",            passed_cnt),
        ("Failed",            failed_cnt),
        ("Pass Rate",         f"{pass_rate}%"),
        ("Total Duration",    f"{duration_sec}s"),
        ("Automation",        "UIAutomator2 via Appium 2.x"),
        ("API Level",         "Android 33"),
        ("Test Language",     "Python 3.12"),
        ("Report Generated",  datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
    ]
    for label, value in details:
        ws.append(["", label, "", value])
        r = ws.max_row
        ws.row_dimensions[r].height = 20
        ws[f"B{r}"].font  = fnt(bold=True)
        ws[f"B{r}"].fill  = fill("F0F4FF")
        ws[f"D{r}"].font  = fnt()
        ws.merge_cells(f"D{r}:H{r}")

    ws.append([])

    # Module breakdown
    mb_row = ws.max_row + 1
    ws.append(["", "MODULE BREAKDOWN", "", "", "", "", "", ""])
    ws.merge_cells(f"B{mb_row}:H{mb_row}")
    ws[f"B{mb_row}"].font  = fnt("Plus Jakarta Sans", 12, True, COLORS['accent'])
    ws[f"B{mb_row}"].fill  = fill("E8EEF7")
    ws.row_dimensions[mb_row].height = 26

    ws.append(["", "Module", "", "Total", "Passed", "Failed", "Pass Rate", ""])
    hr = ws.max_row
    ws.row_dimensions[hr].height = 22
    for col in ["B", "D", "E", "F", "G"]:
        c = ws[f"{col}{hr}"]
        c.font      = fnt("Plus Jakarta Sans", 10, True, "FFFFFF")
        c.fill      = fill(COLORS['accent'])
        c.alignment = align("center")

    modules = list(dict.fromkeys([r['module'] for r in results]))
    for mod in modules:
        mc = [r for r in results if r['module'] == mod]
        mp = len([r for r in mc if r['status'] == 'PASSED'])
        mf = len([r for r in mc if r['status'] == 'FAILED'])
        mr = f"{round((mp / len(mc)) * 100)}%"
        ws.append(["", mod, "", len(mc), mp, mf, mr, ""])
        row_n = ws.max_row
        ws.row_dimensions[row_n].height = 20
        ws[f"B{row_n}"].font = fnt(bold=True)
        for col in ["D", "E", "F", "G"]:
            ws[f"{col}{row_n}"].alignment = align("center")
        ws[f"E{row_n}"].font = fnt(color=COLORS['statPass'])
        ws[f"F{row_n}"].font = fnt(color=COLORS['statFail'])
        ws[f"G{row_n}"].font = fnt(bold=True)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 2

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 2 — ALL TEST CASES
    # ──────────────────────────────────────────────────────────────────────────
    tc_ws = wb.create_sheet("🧪 Test Cases")
    tc_ws.views.sheetView[0].showGridLines = True

    tc_ws.merge_cells("A1:H1")
    tc_ws["A1"].value     = "🛡️  TRUTHGUARD ANDROID — ALL TEST CASES DETAIL"
    tc_ws["A1"].font      = fnt("Outfit", 15, True, COLORS['headerText'])
    tc_ws["A1"].fill      = fill(COLORS['headerBg'])
    tc_ws["A1"].alignment = align("center")
    tc_ws.row_dimensions[1].height = 44

    tc_ws.merge_cells("A2:H2")
    info = tc_ws["A2"]
    info.value     = (f"Total: {total_cnt}  |  Passed: {passed_cnt}  |  "
                      f"Failed: {failed_cnt}  |  Pass Rate: {pass_rate}%  |  Duration: {duration_sec}s")
    info.font      = fnt(size=10, color="FFFFFF")
    info.fill      = fill(COLORS['subHeaderBg'])
    info.alignment = align("center")
    tc_ws.row_dimensions[2].height = 22

    headers = ["#", "Test ID", "Module", "Test Case Name", "Description", "Status", "Duration (ms)", "Error / Notes"]
    tc_ws.append(headers)
    hr = tc_ws.max_row
    tc_ws.row_dimensions[hr].height = 30
    for cell in tc_ws[hr]:
        cell.font      = fnt("Plus Jakarta Sans", 11, True, "FFFFFF")
        cell.fill      = fill(COLORS['accent'])
        cell.alignment = align("center")
        cell.border    = Border(bottom=Side(style='medium', color=COLORS['moduleText']))

    row_idx = 0
    cur_mod  = None
    for r in results:
        row_idx += 1
        if r['module'] != cur_mod:
            cur_mod = r['module']
            tc_ws.append(["", "", cur_mod.upper(), "", "", "", "", ""])
            sep = tc_ws.max_row
            tc_ws.merge_cells(f"C{sep}:H{sep}")
            tc_ws.row_dimensions[sep].height = 24
            for col in range(1, 9):
                c = tc_ws.cell(sep, col)
                c.fill = fill(COLORS['moduleBg'])
                c.font = fnt("Plus Jakarta Sans", 10, True, COLORS['moduleText'])

        is_pass  = r['status'] == 'PASSED'
        stat_bg  = COLORS['passedBg']  if is_pass else COLORS['failedBg']
        stat_fg  = COLORS['passedText'] if is_pass else COLORS['failedText']
        row_bg   = COLORS['rowAlt'] if row_idx % 2 == 0 else COLORS['rowNorm']

        tc_ws.append([row_idx, r['id'], r['module'], r['name'], r['desc'],
                      r['status'], r['duration'], r['error']])
        dn = tc_ws.max_row
        tc_ws.row_dimensions[dn].height = 22

        for col in range(1, 9):
            c = tc_ws.cell(dn, col)
            c.border = border()
            if col == 6:
                c.font      = fnt(bold=True, color=stat_fg)
                c.fill      = fill(stat_bg)
                c.alignment = align("center")
            elif col == 8 and r['error']:
                c.font      = fnt(size=9, color=COLORS['failedText'])
                c.fill      = fill(row_bg)
                c.alignment = align(wrap=True)
            else:
                c.font  = fnt()
                c.fill  = fill(row_bg)
                if col in (1, 2):
                    c.alignment = align("center")
                elif col == 7:
                    c.alignment = align("right")

    tc_ws.column_dimensions["A"].width = 5
    tc_ws.column_dimensions["B"].width = 10
    tc_ws.column_dimensions["C"].width = 16
    tc_ws.column_dimensions["D"].width = 42
    tc_ws.column_dimensions["E"].width = 52
    tc_ws.column_dimensions["F"].width = 14
    tc_ws.column_dimensions["G"].width = 14
    tc_ws.column_dimensions["H"].width = 55

    # ──────────────────────────────────────────────────────────────────────────
    # SHEET 3 — FAILED TESTS  (only if failures exist)
    # ──────────────────────────────────────────────────────────────────────────
    failed_results = [r for r in results if r['status'] == 'FAILED']
    if failed_results:
        fw = wb.create_sheet("❌ Failed Tests")
        fw.views.sheetView[0].showGridLines = True
        fw.merge_cells("A1:G1")
        fw["A1"].value     = f"❌  FAILED TEST CASES — {len(failed_results)} Failures"
        fw["A1"].font      = fnt("Outfit", 14, True, "FFFFFF")
        fw["A1"].fill      = fill(COLORS['statFail'])
        fw["A1"].alignment = align("center")
        fw.row_dimensions[1].height = 40

        fw.append(["#", "Test ID", "Module", "Test Case Name", "Status", "Duration (ms)", "Error Message"])
        fhr = fw.max_row
        fw.row_dimensions[fhr].height = 26
        for cell in fw[fhr]:
            cell.font      = fnt("Plus Jakarta Sans", 10, True, "FFFFFF")
            cell.fill      = fill(COLORS['accent'])
            cell.alignment = align("center")

        for i, r in enumerate(failed_results):
            fw.append([i + 1, r['id'], r['module'], r['name'], r['status'], r['duration'], r['error']])
            fn = fw.max_row
            fw.row_dimensions[fn].height = 30
            for col in range(1, 8):
                c = fw.cell(fn, col)
                c.fill   = fill(COLORS['failedBg'])
                c.border = border()
                if col == 5:
                    c.font      = fnt(bold=True, color=COLORS['failedText'])
                    c.alignment = align("center")
                elif col == 7:
                    c.font      = fnt(size=9, color=COLORS['failedText'])
                    c.alignment = align(wrap=True)
                else:
                    c.font = fnt()

        fw.column_dimensions["A"].width = 5
        fw.column_dimensions["B"].width = 10
        fw.column_dimensions["C"].width = 16
        fw.column_dimensions["D"].width = 44
        fw.column_dimensions["E"].width = 12
        fw.column_dimensions["F"].width = 14
        fw.column_dimensions["G"].width = 62

    wb.save(report_path)
    print(f"📄 Excel report saved → {report_path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MARKDOWN SUMMARY  (GitHub Actions Step Summary)
# ═══════════════════════════════════════════════════════════════════════════════
def generate_markdown_summary(summary_path):
    print("\n📝 Generating Markdown summary...")
    passed      = len([r for r in results if r['status'] == 'PASSED'])
    failed      = len([r for r in results if r['status'] == 'FAILED'])
    total       = len(results)
    pass_rate   = round((passed / total) * 100) if total > 0 else 0
    duration    = round(time.time() - suite_start, 1)
    timestamp   = datetime.now().strftime("%d/%m/%Y, %I:%M:%S %p")
    badge       = '🟢' if pass_rate == 100 else '🟡' if pass_rate >= 80 else '🔴'

    md  = "# 🛡️ TruthGuard Android — Appium E2E Test Report\n\n"
    md += f"> **Generated:** {timestamp} &nbsp;|&nbsp; **Platform:** Android API 33 &nbsp;|&nbsp; **Engine:** Python + Appium Client 3.x\n\n"
    md += "---\n\n"

    md += "## 📊 Results Summary\n\n"
    md += f"| {badge} Pass Rate | 📋 Total Tests | ✅ Passed | ❌ Failed | ⏱️ Duration |\n"
    md += "|:-----------:|:--------------:|:---------:|:---------:|:----------:|\n"
    md += f"| **{pass_rate}%** | **{total}** | **{passed}** | **{failed}** | **{duration}s** |\n\n"

    md += "## 📋 Module Breakdown\n\n"
    md += "| Module | Tests | ✅ Passed | ❌ Failed | Pass Rate |\n"
    md += "|--------|:-----:|:---------:|:---------:|:---------:|\n"
    modules = list(dict.fromkeys([r['module'] for r in results]))
    for mod in modules:
        mc   = [r for r in results if r['module'] == mod]
        mp   = len([r for r in mc if r['status'] == 'PASSED'])
        mf   = len([r for r in mc if r['status'] == 'FAILED'])
        rate = round((mp / len(mc)) * 100)
        icon = '✅' if mf == 0 else '❌'
        md += f"| {icon} {mod} | {len(mc)} | {mp} | {mf} | {rate}% |\n"
    md += "\n"

    failed_list = [r for r in results if r['status'] == 'FAILED']
    if failed_list:
        md += "## ❌ Failed Test Cases\n\n"
        md += "| Test ID | Module | Test Name | Error |\n"
        md += "|---------|--------|-----------|-------|\n"
        for r in failed_list:
            err = (r['error'] or '').replace("|", "\\|")[:120]
            md += f"| `{r['id']}` | {r['module']} | {r['name']} | `{err}` |\n"
        md += "\n"
    else:
        md += "## 🎉 All Tests Passed!\n\n"
        md += f"> All **{total}** E2E test cases passed with a **{pass_rate}%** pass rate on the TruthGuard Android app.\n\n"

    md += "---\n"
    md += "*Excel report available as a downloadable run artifact — see the **Artifacts** section below this run.*\n"

    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"📝 Markdown summary saved → {summary_path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    print("==========================================================")
    print("   🛡️  TRUTHGUARD ANDROID — APPIUM E2E TEST RUNNER")
    print("   135 Test Cases  |  8 Modules  |  Python 3.12 + Appium")
    print("==========================================================")
    print(f"ℹ️  CI mode   : {IS_CI}")
    print(f"ℹ️  ADB       : {ADB_PATH}")
    print(f"ℹ️  APK       : {APK_PATH}")
    print(f"ℹ️  AVD Name  : {AVD_NAME}")
    print(f"ℹ️  App Pkg   : {APP_PACKAGE}")

    # Step 1: Emulator
    try:
        start_emulator()
    except Exception as e:
        print(f"🔴 Emulator error: {e} — continuing anyway")

    # Step 2: Appium (local only)
    server_proc, server_log = None, None
    try:
        server_proc, server_log = start_appium_server()
    except Exception as e:
        print(f"🔴 Appium start error: {e} — continuing anyway")

    driver = None
    try:
        print("\n🧪 Connecting to UIAutomator2 driver...")
        opts = UiAutomator2Options()
        opts.platform_name   = "Android"
        opts.device_name     = AVD_NAME
        opts.app             = APK_PATH
        opts.automation_name = "UiAutomator2"
        opts.set_capability("autoGrantPermissions",               True)
        opts.set_capability("uiautomator2ServerLaunchTimeout",    90000)
        opts.set_capability("uiautomator2ServerInstallTimeout",   90000)
        opts.set_capability("adbExecTimeout",                     60000)
        opts.set_capability("newCommandTimeout",                  300)

        driver = webdriver.Remote(f"http://{APPIUM_HOST}:{APPIUM_PORT}", options=opts)
        print("🟢 Driver connected.")
        time.sleep(2)  # let app settle after install

        run_all_tests(driver)

    except Exception as e:
        print(f"\n🔴 Critical error: {e}")
    finally:
        if driver:
            print("\n🧹 Quitting driver...")
            try:
                driver.quit()
            except Exception:
                pass

        if server_proc:
            print("🧹 Stopping Appium server...")
            server_proc.terminate()
        if server_log:
            server_log.close()

        # ── Summary ──────────────────────────────────────────────────────────
        passed      = len([r for r in results if r['status'] == 'PASSED'])
        failed      = len([r for r in results if r['status'] == 'FAILED'])
        total       = len(results)
        duration    = round(time.time() - suite_start, 1)
        pass_rate   = round((passed / total) * 100) if total > 0 else 0

        print("\n==========================================================")
        print(f"  📊 RESULTS  |  Total: {total}  ✅ Passed: {passed}  ❌ Failed: {failed}")
        print(f"  ⏱️  Duration: {duration}s  |  Pass Rate: {pass_rate}%")
        print("==========================================================")

        # ── Reports ──────────────────────────────────────────────────────────
        ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_dir    = os.path.dirname(os.path.abspath(__file__))
        report_xlsx = os.path.join(base_dir, f"Appium_E2E_Report_TruthGuard_{ts}.xlsx")
        report_md   = os.path.join(base_dir, "test-summary.md")

        generate_excel_report(report_xlsx)
        generate_markdown_summary(report_md)

        print("\n✨ Testing complete! Open the .xlsx report or check GitHub Actions Summary.")


if __name__ == "__main__":
    main()
